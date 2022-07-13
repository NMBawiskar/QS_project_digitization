import os
from configparser import ConfigParser
# import hashlib
import datetime
import json
from openpyxl import Workbook

from flask import Flask, request, jsonify
from flask_jwt_extended import (
    JWTManager, jwt_required, create_access_token,
    get_jwt_identity
)
from flask_mysqldb import MySQL
from flask_swagger import swagger
from flask_swagger_ui import get_swaggerui_blueprint

import database_credentials
import bcrypt
from werkzeug.utils import secure_filename

# os.chdir('/home/mdindia/Digitization')

configuration = ConfigParser()
# print(os.getcwd())
configuration.read('config.ini')
app = Flask(__name__)
app.config['MYSQL_HOST'] = database_credentials.host
app.config['MYSQL_USER'] = database_credentials.user
app.config['MYSQL_PASSWORD'] = database_credentials.password
app.config['MYSQL_DB'] = database_credentials.database
app.config['MYSQL_PORT'] = database_credentials.port
app.config['MYSQL_CONNECT_TIMEOUT'] = 28800

app.config['JWT_SECRET_KEY'] = 'SecurityIsAnIllusion'
jwt = JWTManager(app)

app.secret_key = 'SecurityIsAnIllusion'
mysql_connector = MySQL(app)

SWAGGER_URL = configuration.get('swagger', 'swagger_url')
API_URL = configuration.get('swagger', 'api_url')
# SWAGGER_URL = "/api-docs"
# API_URL = "/api/docs"
ALLOWED_FILE_EXTENSIONS = ['pdf']

# TODO - Change these SPs for the new server
# reports_get_status = "CALL reports_get_status('{ids}');"

# These have been changed
users_login = "CALL system_user_get_pass('{username}');"
report_upsert = "SET @x = {rec_id}; CALL reports_upsert(@x, '{proposal_id}', {vendor_id}, '{life_assured_of}', '{report_name}', {page_count}, '{status}'); SELECT @x;"
report_progress_log = "SET @rx = 0; CALL report_progress_insert(@rx, {report_id}, '{status}', {user_id}, '{time_taken}', '{comment}');"
reports_get_status_by_proposal_id = "CALL reports_get_status_by_proposal_id('{proposal_ids}');"


def exec_query(sql, params):
    mysql_connector.connection.autocommit(1)
    sql = sql.format(**params)
    print(sql)
    mysql_response = []
    with mysql_connector.connection.cursor() as cursor:
        for statement in sql.split(';'):
            if len(statement.strip()) > 0:
                cursor.execute(statement.strip() + ';')
                mysql_response.append(cursor.fetchall())
    cursor.close()
    return mysql_response


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[-1].lower() in ALLOWED_FILE_EXTENSIONS


@app.route('/vendor/login', methods=['POST'])
def login():
    """
        Login API for authentication.
        ---
        tags:
          - Authentication
        summary: Login API for authentication.
        description: The API takes in a username-password pair and returns an `access_token` that must be passed in the headers of all other APIs.
        definitions:
          - schema:
              id: Login
              properties:
                username:
                 type: string
                password:
                 type: string
          - schema:
              id: Status_response
              type: array
              items:
                type: object
                properties:
                  file_id:
                    type: number
                    required: true
                  status:
                    type: string
                    required: true
                    enum:
                      - In Queue
                      - Under Processing
                      - Digitized
                      - Under QC
                      - Completed
                      - Issue
                      - Faulty
                  file:
                    type: string
                    description: A link to the digitized file in the required format, *if* the file has been processed and verified by QC.
        parameters:
          - in: body
            name: Login body
            description: A username-password pair for `access_token` generation.
            required: true
            schema:
               $ref: "#/definitions/Login"
        responses:
          200:
            description: Returns a bearer token that must be passed in the headers of all other APIs.
            schema:
              properties:
                access_token:
                 type: string
          400:
            description: Missing parameters in the request body.
            schema:
              properties:
                message:
                 type: string
          401:
            description: Invalid credentials.
            schema:
              properties:
                message:
                 type: string
    """
    if not request.is_json:
        return jsonify({"message": "Missing JSON in request"}), 400

    username = request.json.get('username', None)
    password = request.json.get('password', None)
    if not username:
        return jsonify({"message": "Missing username parameter"}), 400
    if not password:
        return jsonify({"message": "Missing password parameter"}), 400

    response = exec_query(users_login, {'username': username})
    if len(response[0]) != 1:
        return jsonify({"message": "Bad username or password"}), 401
    elif not bcrypt.checkpw(password.encode('utf8'), response[0][0][2].encode('utf8')) and response[0][0][1] == 'Vendor':
        return jsonify({"message": "Bad username or password"}), 401

    # Identity can be any data that is json serializable
    access_token = create_access_token(identity=response[0][0][0], expires_delta=datetime.timedelta(hours=24))
    return jsonify(access_token=access_token), 200

@jwt_required
@app.route('/v1/upload-file', methods=['POST'])
def upload_file():
    """
            Upload a report for Digitization.
            ---
            tags:
              - Digitization
            summary: Upload a report for Digitization.
            consumes: multipart/form-data
            parameters:
              - in: formData
                name: file
                description: PDF report that needs to be digitized.
                type: file
                required: true
              - in: formData
                name: proposal_id
                description: Proposal ID for the case specified.
                type: string
                required: true
              - in: formData
                name: vendor_id
                description: Vendor ID of the vendor from whom this report comes from.
                type: string
                required: true
              - in: formData
                name: life_assured_of
                description: Name of the life assured.
                type: string
            responses:
              200:
                description: The `report_id` of the uploaded file. This `report_id` can be used to get the progress status of the particular report.
                schema:
                  properties:
                    message:
                      type: string
                    report_id:
                      type: number
              204:
                description: Missing file or invalid file format.
                schema:
                  properties:
                    message:
                      type: string
            security:
              - Bearer: []
        """
    if 'file' in request.files:
        file = request.files['file']
        meta = request.form
        if allowed_file(file.filename):
            try:
                # hash_file_name = hashlib.md5(file.read()).hexdigest() + '.' + file.filename.rsplit('.', 1)[-1].lower()
                hash_file_name = secure_filename(file.filename)
                path_to_reports = os.path.join(os.getcwd(), configuration.get('digitization', 'default_reporting_directory'))
                path_to_input_folder = os.path.join(path_to_reports, configuration.get('digitization', 'input_path'))
                os.makedirs(path_to_reports, exist_ok=True)
                os.makedirs(path_to_input_folder, exist_ok=True)
                file.save(os.path.join(path_to_input_folder, hash_file_name))
                res = exec_query(report_upsert, {'rec_id': 0, 'proposal_id': meta['proposal_id'], 'vendor_id': meta['vendor_id'], 'life_assured_of': meta['life_assured_of'] if 'life_assured_of' in meta.keys() else '', 'report_name': hash_file_name.rsplit('.', 1)[0], 'page_count': 0, 'status': 'In Queue'})
                exec_query(report_progress_log, {'report_id': res[2][0][0], 'status': 'In Queue', 'user_id': int(get_jwt_identity()), 'time_taken': '', 'comment': ''})
                return jsonify({'message': 'Successfully added', 'report_id': res[2][0][0]}), 200
            except Exception as e:
                print(e)
                return jsonify({'message': 'Something went wrong. Please contact our system administrator.'}), 200
        else:
            return jsonify(message='Invalid file format. Please upload one of the following formats: ' + ', '.join(ALLOWED_FILE_EXTENSIONS)), 200
    else:
        return jsonify(message='Please attach a file in the body.'), 200

@jwt_required
@app.route('/v1/get-file-status', methods=['POST'])
def get_file_status():
    """
      Get the status for uploaded files by `proposal_id`.
      ---
      tags:
        - Digitization
      summary: Get the status for uploaded files by `proposal_id`.
      parameters:
        - in: body
          name: proposal_ids
          description: An array of comma-separated `proposal_id` of the reports whose status is needed.
          type: array
          items:
              type: string
          required: true
      responses:
        200:
          description: Status of the files matching the uploaded file IDs.
          schema:
            $ref: "#/definitions/Status_response"
        204:
          description: Missing parameters in the request body.
          schema:
            properties:
              message:
                type: string
      security:
        - Bearer: []
      """
    if type(request.get_json()) == list:
        res = exec_query(reports_get_status_by_proposal_id, {'proposal_ids': ','.join([str(i) for i in request.get_json()])})
        file_data = []
        for t in res[0]:
            obj = {'file_id': t[0], 'status': t[4]}
            if obj['status'] == 'Completed':
                with open(os.path.join('Reports/extracted_json_files', t[1] + ".json"), 'r') as f:
                    if t[2] == 61:
                        create_excel_for_max_life(json.loads(f.read()), t[1], t[3])
            file_data.append(obj)
        return jsonify(file_data), 200
    else:
        return jsonify(message='Please send the body in the specified format'), 204


@app.route("/api/docs")
def spec():
    swag = swagger(app)
    swag['info']['version'] = "1.7"
    swag['info']['title'] = "Digitization APIs"
    swag['info']['description'] = """
        This is the API documentation for QuicSolv's Digitization. 
        Steps for authenticating the APIs:
        1. Login using the login API. The response will contain an `access_token` that will be used for further authentication.
        2. Use this token to authenticate all other APIs by adding it in the Authorization header as `Bearer {access_token}`.
           For swagger, you can simply click on the 'Authorize' button and set it as `Bearer {access_token}`.
    """
    swag['securityDefinitions'] = {"Bearer": {"type": "apiKey", "scheme": "bearer", "name": "Authorization", "in": "header"}}
    return jsonify(swag)


def create_excel_for_max_life(data, filename, proposal_id):
    wb = Workbook()
    ws = wb.active
    cols = ['proposalNo', 'insuredName', 'proposerName', 'isGenderfemale',
            'dob', 'contactNo', 'email', 'gender',
            'fbsResult', 'fbsMin', 'fbsMax',
            'sCholesterolResult', 'sCholesterolMin', 'sCholesterolMax',
            'hdlResult', 'hdlMin', 'hdlMax',
            'triglyceridesResult', 'triglyceridesMin', 'triglyceridesMax',
            'sCreatinineResult', 'sCreatinineMin', 'sCreatinineMax',
            'sAlbuminResult', 'sAlbuminMin', 'sAlbuminMax',
            'sGOTResult', 'sGOTMin', 'sGOTMax',
            'sGPTResult', 'sGPTMin', 'sGPTMax',
            'hIV1and2Result', 'hBsAgResult',
            'alkalineResult', 'alkalineMin', 'alkalineMax',
            'bilirubinResult', 'bilirubinMin', 'bilirubinMax',
            'ggtResult', 'ggtMin', 'ggtmax',
            'bunResult', 'bunMin', 'bunMax',
            'albuminProtienResult', 'glucoseSugarResult', 'pusCellResult',
            'urine RBCResult', 'urineWBCResult', 'ketoneResult',
            'haemoglobinResult', 'rbcResult', 'wbcResult',
            'mcvResult', 'plateletsResult', 'esrResult',
            'ecgResult', 'tmtResult', 'urineCotinineResult',
            'hbA1CResult', 'hbA1CMin', 'hbA1CMax',
            'ppbs Result', 'ppbsSMin', 'ppbsMax']

    # print("Cols:", len(cols))
    # print(data['TESTS'])

    found_items = {}
    for _, values in data['TESTS'].items():
        for item in values:
            if 'TST' in item:
                if item['TST'].lower() in ['haemoglobin', 'rbc', 'wbc', 'platelet', 'mcv']:
                    found_items[item['TST'].lower()] = item

    col_data = dict()
    col_data[0] = proposal_id
    col_data[1] = str(next(iter(data['PER'])))
    col_data[3] = str(next(iter(data['SEX']))) if 'F' in str(next(iter(data['SEX']))).upper() else 'No'
    col_data[7] = str(next(iter(data['SEX'])))

    for idx, row in enumerate(ws.iter_rows(min_row=1, max_col=len(cols), max_row=2)):
        if idx == 0:
            for jdx, cell in enumerate(row):
                cell.value = cols[jdx]
        else:
            ws['A2'] = proposal_id
            for jdx, cell in enumerate(row):
                cell.value = col_data[jdx].upper() if jdx in col_data.keys() else ''
                test_name = cols[jdx].replace('Result', '').replace('Min', '').replace('Max', '')
                # print(jdx + 1, test_name)
                if test_name in found_items.keys():
                    sub_val = cols[jdx].replace(test_name, '')
                    if sub_val == 'Result':
                        cell.value = found_items[test_name]['RES']
                    elif sub_val == 'Min' and len(found_items[test_name]['REF'].split('-')) >= 1:
                        cell.value = found_items[test_name]['REF'].split('-')[0]
                    elif sub_val == 'Max' and len(found_items[test_name]['REF'].split('-')) == 2:
                        cell.value = found_items[test_name]['REF'].split('-')[-1]

    wb.save(os.path.join('Reports/extracted_excel_files', filename + '.xlsx'))


swaggerui_blueprint = get_swaggerui_blueprint(
        SWAGGER_URL,  # Swagger UI static files will be mapped to '{SWAGGER_URL}/dist/'
        API_URL,
        config={  # Swagger UI config overrides
            'app_name': "Digitization"
        },
    )

app.register_blueprint(swaggerui_blueprint)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=2000, debug=True)
